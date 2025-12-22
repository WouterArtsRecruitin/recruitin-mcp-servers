#!/usr/bin/env python3
"""
CV-VACANCY MATCHER MCP SERVER
Recruitin B.V. - Wouter Arts

Allows Claude/AI to directly match candidates with vacancies via natural language.

Usage:
  # Development
  fastmcp dev server.py
  
  # Production  
  fastmcp run server.py
  
  # Claude Desktop config:
  {
    "mcpServers": {
      "cv-matcher": {
        "command": "fastmcp",
        "args": ["run", "/path/to/server.py"],
        "env": {
          "HF_TOKEN": "hf_...",
          "RESEND_API_KEY": "re_..."
        }
      }
    }
  }
"""

import os
import json
import csv
import urllib.request
import base64
from dataclasses import dataclass, field, asdict
from typing import List, Optional
from datetime import datetime
from fastmcp import FastMCP

# ============================================
# CONFIGURATION
# ============================================

CONFIG = {
    "hf_token": os.environ.get("HF_TOKEN", ""),
    "hf_model": "sentence-transformers/all-MiniLM-L6-v2",
    "resend_api_key": os.environ.get("RESEND_API_KEY", ""),
    "from_email": "CV Matcher <onboarding@resend.dev>",
    "reply_to": "warts@recruitin.nl",
    "typeform_url": "https://form.typeform.com/to/uwu2PZyR",
    "vacatures_path": os.environ.get("VACATURES_CSV", "vacatures.csv"),
}

# ============================================
# DATA CLASSES
# ============================================

@dataclass
class Kandidaat:
    voornaam: str
    achternaam: str
    email: str = ""
    gewenste_functie: str = ""
    jaren_ervaring: int = 0
    skills: List[str] = field(default_factory=list)
    locatie: str = ""
    
    @property
    def volledige_naam(self): 
        return f"{self.voornaam} {self.achternaam}"
    
    def search_text(self):
        return f"{self.gewenste_functie} {' '.join(self.skills)} {self.locatie}"

@dataclass 
class Vacature:
    id: str
    titel: str
    bedrijf: str
    locatie: str = ""
    skills: List[str] = field(default_factory=list)
    salaris_min: int = 0
    salaris_max: int = 0
    contact_naam: str = ""
    contact_email: str = ""
    vacature_url: str = ""
    
    def search_text(self):
        return f"{self.titel} {self.bedrijf} {' '.join(self.skills)} {self.locatie}"

@dataclass
class Match:
    vacature: Vacature
    score: float
    ai_score: float

# ============================================
# CORE FUNCTIONS
# ============================================

def load_vacatures(path: str = None) -> List[Vacature]:
    """Load vacatures from CSV"""
    path = path or CONFIG["vacatures_path"]
    
    if not os.path.exists(path):
        # Return sample data if no CSV
        return [
            Vacature("1", "Ploegleider Productie", "ASML", "Veldhoven", 
                    ["leiderschap","productie","lean"], 55000, 70000,
                    "Sandra Vermeulen", "s.vermeulen@asml.com", "https://asml.com/jobs/1"),
            Vacature("2", "Teamleider Assemblage", "Philips", "Eindhoven",
                    ["teamleider","assemblage","kwaliteit"], 50000, 65000,
                    "Marc de Jong", "m.dejong@philips.com", "https://philips.com/jobs/2"),
            Vacature("3", "Production Supervisor", "VDL", "Eindhoven",
                    ["supervisor","manufacturing"], 48000, 60000,
                    "Lisa Peters", "l.peters@vdl.nl", "https://vdl.nl/jobs/3"),
            Vacature("4", "Process Engineer", "ASML", "Veldhoven",
                    ["engineer","lean","six-sigma"], 60000, 80000,
                    "Tom Bakker", "t.bakker@asml.com", "https://asml.com/jobs/4"),
            Vacature("5", "Operations Manager", "Vanderlande", "Veghel",
                    ["operations","lean","management"], 65000, 85000,
                    "Erik Mulder", "e.mulder@vanderlande.com", "https://vanderlande.com/jobs/5"),
        ]
    
    vacatures = []
    with open(path, encoding="utf-8") as f:
        for row in csv.DictReader(f, delimiter=";"):
            skills = [s.strip() for s in row.get("skills","").split(",") if s.strip()]
            vacatures.append(Vacature(
                id=row.get("id",""),
                titel=row.get("functietitel",""),
                bedrijf=row.get("bedrijfsnaam",""),
                locatie=row.get("locatie",""),
                skills=skills,
                salaris_min=int(row.get("salaris_min",0) or 0),
                salaris_max=int(row.get("salaris_max",0) or 0),
                contact_naam=f"{row.get('contact_voornaam','')} {row.get('contact_achternaam','')}".strip(),
                contact_email=row.get("contact_email",""),
                vacature_url=row.get("vacature_url",""),
            ))
    return vacatures

def get_embeddings(texts: List[str]) -> List[List[float]]:
    """Get embeddings from HuggingFace"""
    token = CONFIG["hf_token"]
    if not token:
        return []
    
    try:
        from huggingface_hub import InferenceClient
        client = InferenceClient(token=token)
        result = client.feature_extraction(texts, model=CONFIG["hf_model"])
        if hasattr(result[0], 'tolist'):
            return [r.tolist() for r in result]
        return list(result)
    except Exception as e:
        print(f"HF error: {e}")
        return []

def cosine_sim(a: List[float], b: List[float]) -> float:
    """Cosine similarity between two vectors"""
    dot = sum(x*y for x,y in zip(a,b))
    na = sum(x*x for x in a)**0.5
    nb = sum(x*x for x in b)**0.5
    return dot/(na*nb) if na and nb else 0

def match_kandidaat_vacatures(kandidaat: Kandidaat, vacatures: List[Vacature]) -> List[Match]:
    """AI-powered matching"""
    texts = [kandidaat.search_text()] + [v.search_text() for v in vacatures]
    embs = get_embeddings(texts)
    
    results = []
    
    if len(embs) >= 2:
        # AI matching
        k_emb = embs[0]
        for i, v in enumerate(vacatures):
            ai = (cosine_sim(k_emb, embs[i+1]) + 1) / 2
            
            # Keyword bonus
            kw = 0
            if kandidaat.gewenste_functie.lower() in v.titel.lower():
                kw += 0.3
            skill_match = len(set(s.lower() for s in kandidaat.skills) & 
                           set(s.lower() for s in v.skills))
            kw += skill_match * 0.1
            
            score = ai * 0.6 + min(kw, 0.4)
            results.append(Match(v, min(score, 1.0), ai))
    else:
        # Fallback: keyword matching
        for v in vacatures:
            score = 0
            if kandidaat.gewenste_functie.lower() in v.titel.lower():
                score += 0.5
            skills_k = set(s.lower() for s in kandidaat.skills)
            skills_v = set(s.lower() for s in v.skills)
            if skills_v:
                score += len(skills_k & skills_v) / len(skills_v) * 0.5
            results.append(Match(v, score, 0))
    
    results.sort(key=lambda x: x.score, reverse=True)
    return results

def generate_excel(kandidaat: Kandidaat, matches: List[Match]) -> str:
    """Generate Excel report"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return ""
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Matches"
    
    headers = ["#", "Score", "Vacature", "Bedrijf", "Locatie", "Contact", "Email"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(1, c, h)
        cell.fill = PatternFill("solid", fgColor="1E3A5F")
        cell.font = Font(color="FFFFFF", bold=True)
    
    for i, m in enumerate(matches[:10], 1):
        ws.cell(i+1, 1, i)
        ws.cell(i+1, 2, f"{int(m.score*100)}%")
        ws.cell(i+1, 3, m.vacature.titel)
        ws.cell(i+1, 4, m.vacature.bedrijf)
        ws.cell(i+1, 5, m.vacature.locatie)
        ws.cell(i+1, 6, m.vacature.contact_naam)
        ws.cell(i+1, 7, m.vacature.contact_email)
    
    path = f"/tmp/Matches_{kandidaat.achternaam}_{datetime.now():%Y%m%d_%H%M%S}.xlsx"
    wb.save(path)
    return path

def send_match_email(kandidaat: Kandidaat, matches: List[Match], excel_path: str = None) -> dict:
    """Send email with matches"""
    key = CONFIG["resend_api_key"]
    if not key:
        return {"success": False, "error": "No RESEND_API_KEY"}
    
    if not kandidaat.email:
        return {"success": False, "error": "No email address"}
    
    # Build HTML
    rows = ""
    for i, m in enumerate(matches[:5], 1):
        v = m.vacature
        rows += f"""<tr>
<td style="padding:8px;text-align:center"><b>{i}</b></td>
<td style="padding:8px;text-align:center"><span style="background:#10b981;color:#fff;padding:3px 8px;border-radius:4px">{int(m.score*100)}%</span></td>
<td style="padding:8px"><b>{v.titel}</b></td>
<td style="padding:8px">{v.bedrijf}</td>
<td style="padding:8px">{v.locatie}</td>
<td style="padding:8px">{v.contact_naam}<br><a href="mailto:{v.contact_email}">{v.contact_email}</a></td>
</tr>"""
    
    html = f"""<!DOCTYPE html>
<html><body style="font-family:system-ui;padding:20px">
<div style="max-width:700px;margin:auto;background:#fff;border-radius:12px;overflow:hidden;box-shadow:0 2px 8px rgba(0,0,0,0.1)">
<div style="background:linear-gradient(135deg,#1e3a5f,#2563eb);padding:24px;color:#fff">
<h1 style="margin:0">🎯 CV-Vacature AI Matching</h1>
<p style="margin:8px 0 0">Top vacatures voor {kandidaat.volledige_naam}</p>
</div>
<div style="padding:24px">
<p>Hoi {kandidaat.voornaam},</p>
<p>Op basis van je profiel vonden we <b>{len(matches)} matches</b>!</p>
<table style="width:100%;border-collapse:collapse">
<tr style="background:#f8f9fa"><th>#</th><th>Score</th><th>Vacature</th><th>Bedrijf</th><th>Locatie</th><th>Contact</th></tr>
{rows}
</table>
</div>
<div style="background:#f0f9ff;margin:20px;padding:20px;border-radius:8px;text-align:center">
<a href="{CONFIG['typeform_url']}" style="background:#0ea5e9;color:#fff;padding:12px 24px;border-radius:6px;text-decoration:none">📝 Profiel Bijwerken</a>
</div>
<div style="padding:20px;background:#f8f9fa;border-top:1px solid #eee">
<b>Wouter Arts</b> - Recruitin B.V.<br>
<a href="mailto:warts@recruitin.nl">warts@recruitin.nl</a> | <a href="https://calendly.com/wouter-arts-/15min">Plan gesprek</a>
</div>
</div>
</body></html>"""
    
    data = {
        "from": CONFIG["from_email"],
        "to": [kandidaat.email],
        "reply_to": CONFIG["reply_to"],
        "subject": f"🎯 Top {min(10,len(matches))} vacatures voor {kandidaat.voornaam}",
        "html": html,
    }
    
    if excel_path and os.path.exists(excel_path):
        with open(excel_path, "rb") as f:
            data["attachments"] = [{
                "filename": os.path.basename(excel_path),
                "content": base64.b64encode(f.read()).decode()
            }]
    
    req = urllib.request.Request(
        "https://api.resend.com/emails",
        json.dumps(data).encode(),
        {"Authorization": f"Bearer {key}", "Content-Type": "application/json"}
    )
    
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            res = json.loads(r.read())
            return {"success": True, "id": res.get("id"), "to": kandidaat.email}
    except Exception as e:
        return {"success": False, "error": str(e)}

# ============================================
# MCP SERVER
# ============================================

mcp = FastMCP("cv-vacancy-matcher")

@mcp.tool()
def zoek_vacatures(
    functie: str,
    skills: str = "",
    locatie: str = "",
    max_results: int = 5
) -> str:
    """
    Zoek vacatures die matchen met een functieprofiel.
    
    Args:
        functie: Gewenste functie (bijv. "Ploegleider", "Engineer", "Teamleider")
        skills: Komma-gescheiden skills (bijv. "lean, productie, leiderschap")
        locatie: Gewenste locatie (bijv. "Eindhoven", "Noord-Brabant")
        max_results: Maximum aantal resultaten (default 5)
    
    Returns:
        Lijst van matchende vacatures met scores
    """
    skill_list = [s.strip() for s in skills.split(",") if s.strip()]
    
    kandidaat = Kandidaat(
        voornaam="Zoek",
        achternaam="Query",
        gewenste_functie=functie,
        skills=skill_list,
        locatie=locatie
    )
    
    vacatures = load_vacatures()
    matches = match_kandidaat_vacatures(kandidaat, vacatures)
    top = matches[:max_results]
    
    result = f"🎯 **{len(top)} vacatures gevonden voor '{functie}'**\n\n"
    
    for i, m in enumerate(top, 1):
        v = m.vacature
        result += f"**{i}. {v.titel}** ({int(m.score*100)}% match)\n"
        result += f"   🏢 {v.bedrijf} | 📍 {v.locatie}\n"
        result += f"   💰 €{v.salaris_min:,}-{v.salaris_max:,}\n"
        result += f"   🛠️ {', '.join(v.skills[:5])}\n"
        result += f"   👤 {v.contact_naam} ({v.contact_email})\n"
        if v.vacature_url:
            result += f"   🔗 {v.vacature_url}\n"
        result += "\n"
    
    return result

@mcp.tool()
def match_kandidaat(
    voornaam: str,
    achternaam: str,
    email: str,
    functie: str,
    skills: str = "",
    locatie: str = "",
    jaren_ervaring: int = 0
) -> str:
    """
    Match een specifieke kandidaat met vacatures en genereer rapport.
    
    Args:
        voornaam: Voornaam van de kandidaat
        achternaam: Achternaam van de kandidaat
        email: Email adres van de kandidaat
        functie: Gewenste functie
        skills: Komma-gescheiden skills
        locatie: Gewenste locatie
        jaren_ervaring: Jaren werkervaring
    
    Returns:
        Matching resultaten met top vacatures
    """
    skill_list = [s.strip() for s in skills.split(",") if s.strip()]
    
    kandidaat = Kandidaat(
        voornaam=voornaam,
        achternaam=achternaam,
        email=email,
        gewenste_functie=functie,
        skills=skill_list,
        locatie=locatie,
        jaren_ervaring=jaren_ervaring
    )
    
    vacatures = load_vacatures()
    matches = match_kandidaat_vacatures(kandidaat, vacatures)
    top = matches[:10]
    
    # Generate Excel
    excel_path = generate_excel(kandidaat, top)
    
    result = f"🎯 **Matching voor {kandidaat.volledige_naam}**\n\n"
    result += f"📧 {email}\n"
    result += f"💼 {functie} | {jaren_ervaring} jaar ervaring\n"
    result += f"🛠️ {', '.join(skill_list)}\n"
    result += f"📍 {locatie}\n\n"
    result += f"**Top {len(top)} matches:**\n\n"
    
    for i, m in enumerate(top[:5], 1):
        v = m.vacature
        result += f"{i}. **{v.titel}** @ {v.bedrijf} — **{int(m.score*100)}%**\n"
        result += f"   {v.locatie} | {v.contact_naam} ({v.contact_email})\n"
    
    if len(top) > 5:
        result += f"\n... en {len(top)-5} meer matches\n"
    
    if excel_path:
        result += f"\n📊 Excel rapport: {excel_path}\n"
    
    result += f"\n💡 Gebruik `stuur_matching_email` om resultaten naar {email} te sturen."
    
    return result

@mcp.tool()
def stuur_matching_email(
    voornaam: str,
    achternaam: str,
    email: str,
    functie: str,
    skills: str = "",
    locatie: str = ""
) -> str:
    """
    Match kandidaat en stuur resultaten per email.
    
    Args:
        voornaam: Voornaam van de kandidaat
        achternaam: Achternaam van de kandidaat
        email: Email adres om resultaten naar te sturen
        functie: Gewenste functie
        skills: Komma-gescheiden skills
        locatie: Gewenste locatie
    
    Returns:
        Bevestiging van verzonden email
    """
    skill_list = [s.strip() for s in skills.split(",") if s.strip()]
    
    kandidaat = Kandidaat(
        voornaam=voornaam,
        achternaam=achternaam,
        email=email,
        gewenste_functie=functie,
        skills=skill_list,
        locatie=locatie
    )
    
    vacatures = load_vacatures()
    matches = match_kandidaat_vacatures(kandidaat, vacatures)
    top = matches[:10]
    
    # Generate Excel
    excel_path = generate_excel(kandidaat, top)
    
    # Send email
    result = send_match_email(kandidaat, top, excel_path)
    
    if result["success"]:
        return f"""✅ **Email verzonden!**

📧 Naar: {email}
👤 Kandidaat: {kandidaat.volledige_naam}
🎯 Matches: {len(top)} vacatures
📊 Excel bijlage: ✓

Top 3 matches in email:
1. {top[0].vacature.titel} @ {top[0].vacature.bedrijf} ({int(top[0].score*100)}%)
2. {top[1].vacature.titel} @ {top[1].vacature.bedrijf} ({int(top[1].score*100)}%)
3. {top[2].vacature.titel} @ {top[2].vacature.bedrijf} ({int(top[2].score*100)}%)

Email ID: {result.get('id', 'N/A')}"""
    else:
        return f"❌ **Email verzenden mislukt**\n\nFout: {result.get('error', 'Onbekend')}"

@mcp.tool()
def lijst_vacatures(
    locatie: str = "",
    min_salaris: int = 0
) -> str:
    """
    Toon alle beschikbare vacatures, optioneel gefilterd.
    
    Args:
        locatie: Filter op locatie (optioneel)
        min_salaris: Minimum salaris filter (optioneel)
    
    Returns:
        Lijst van alle vacatures
    """
    vacatures = load_vacatures()
    
    # Apply filters
    if locatie:
        vacatures = [v for v in vacatures if locatie.lower() in v.locatie.lower()]
    if min_salaris > 0:
        vacatures = [v for v in vacatures if v.salaris_min >= min_salaris]
    
    result = f"📋 **{len(vacatures)} vacatures beschikbaar**\n\n"
    
    for v in vacatures:
        result += f"• **{v.titel}** @ {v.bedrijf}\n"
        result += f"  📍 {v.locatie} | 💰 €{v.salaris_min:,}-{v.salaris_max:,}\n"
        result += f"  👤 {v.contact_naam}\n\n"
    
    return result

@mcp.tool()
def vacature_details(vacature_id: str) -> str:
    """
    Toon gedetailleerde informatie over een specifieke vacature.
    
    Args:
        vacature_id: ID van de vacature
    
    Returns:
        Volledige vacature details
    """
    vacatures = load_vacatures()
    
    for v in vacatures:
        if v.id == vacature_id:
            return f"""📋 **{v.titel}**

🏢 **Bedrijf:** {v.bedrijf}
📍 **Locatie:** {v.locatie}
💰 **Salaris:** €{v.salaris_min:,} - €{v.salaris_max:,}

🛠️ **Skills:**
{chr(10).join(f'  • {s}' for s in v.skills)}

👤 **Contact:**
  Naam: {v.contact_naam}
  Email: {v.contact_email}

🔗 **URL:** {v.vacature_url or 'N/A'}"""
    
    return f"❌ Vacature met ID '{vacature_id}' niet gevonden."

@mcp.resource("vacatures://lijst")
def get_vacatures_resource() -> str:
    """Alle vacatures als resource"""
    vacatures = load_vacatures()
    return json.dumps([asdict(v) for v in vacatures], indent=2, default=str)

@mcp.resource("config://status")
def get_config_status() -> str:
    """Server configuratie status"""
    return json.dumps({
        "hf_configured": bool(CONFIG["hf_token"]),
        "resend_configured": bool(CONFIG["resend_api_key"]),
        "hf_model": CONFIG["hf_model"],
        "vacatures_path": CONFIG["vacatures_path"],
        "typeform_url": CONFIG["typeform_url"],
    }, indent=2)

# ============================================
# MAIN
# ============================================

if __name__ == "__main__":
    print("🚀 CV-Vacancy Matcher MCP Server")
    print("=" * 40)
    print(f"HF Token: {'✓' if CONFIG['hf_token'] else '✗'}")
    print(f"Resend Key: {'✓' if CONFIG['resend_api_key'] else '✗'}")
    print(f"Vacatures: {CONFIG['vacatures_path']}")
    print()
    print("Tools beschikbaar:")
    print("  • zoek_vacatures")
    print("  • match_kandidaat")
    print("  • stuur_matching_email")
    print("  • lijst_vacatures")
    print("  • vacature_details")
    print()
    print("Start met: fastmcp dev server.py")
    
    mcp.run()
